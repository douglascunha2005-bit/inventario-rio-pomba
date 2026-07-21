# -*- coding: utf-8 -*-
"""
Gera index.html a partir da planilha .xlsx do repositório.
Regras: 1 aba, sem células mescladas, colunas A-D = níveis 1-4, E = código.
Célula vazia em A-D herda o nível anterior; o código fica no último nível preenchido.
Falha (exit 1) se os dados estiverem inconsistentes — nada errado é publicado.
"""
import base64
import glob
import hashlib
import json
import re
import sys

from openpyxl import load_workbook


def erro(msg):
    print(f"ERRO: {msg}")
    sys.exit(1)


def clean(s):
    if s is None:
        return None
    s = re.sub(r"[\x00-\x1f\x7f\u200b\u200c\u200d\ufeff]", "", str(s)).strip()
    return s[:200] if s else None


# ── 1. Localizar a planilha (exatamente UMA .xlsx no repositório) ──
arquivos = [a for a in glob.glob("*.xlsx") if not a.startswith("~$")]
if len(arquivos) == 0:
    erro("nenhuma planilha .xlsx encontrada no repositório.")
if len(arquivos) > 1:
    erro(f"mais de uma .xlsx no repositório: {arquivos}. Mantenha apenas uma.")
planilha = arquivos[0]
print(f"Planilha: {planilha}")

# ── 2. Ler e validar ──
wb = load_workbook(planilha, data_only=True)
if len(wb.sheetnames) > 1:
    erro(f"a planilha tem {len(wb.sheetnames)} abas ({wb.sheetnames}). Mantenha apenas uma.")
ws = wb[wb.sheetnames[0]]
if len(ws.merged_cells.ranges) > 0:
    erro(f"há células mescladas: {[str(r) for r in ws.merged_cells.ranges]}. Desfaça as mesclagens.")

flat = []
codigos = {}
for i, row in enumerate(ws.iter_rows(min_row=2, max_col=5, values_only=True), 2):
    a, b, c, d, e = row
    vals = [clean(a), clean(b), clean(c), clean(d)]
    code = None
    if e is not None:
        code = str(int(e)) if isinstance(e, (int, float)) else clean(e)
        if code and not code.isdigit():
            erro(f"linha {i}: código não numérico ({code!r}).")
    filled = [(lvl, v) for lvl, v in enumerate(vals) if v]
    if not filled:
        if code:
            erro(f"linha {i}: código {code} sem nome em nenhuma coluna A-D.")
        continue  # linha totalmente vazia — ignora
    if len(filled) > 1:
        erro(f"linha {i}: mais de uma coluna preenchida {filled}. Cada linha deve ter texto em UMA coluna só.")
    lvl, name = filled[-1]
    if code:
        if code in codigos:
            erro(f"código {code} duplicado (linhas {codigos[code]} e {i}).")
        codigos[code] = i
    flat.append((i, lvl, name, code))

# hierarquia: nível N exige um nível N-1 anterior
stack = [False, False, False, False]
for i, lvl, name, code in flat:
    if lvl > 0 and not stack[lvl - 1]:
        erro(f"linha {i}: '{name}' está no nível {lvl+1} sem pai no nível {lvl}.")
    stack[lvl] = True
    for j in range(lvl + 1, 4):
        stack[j] = False

print(f"Validação OK: {len(flat)} nós, {len(codigos)} códigos, sem duplicatas.")

# ── 3. Injetar no template e recalcular CSP ──
data_js = json.dumps([[lvl, name, code] for _, lvl, name, code in flat],
                     ensure_ascii=False, separators=(",", ":"))
html = open("template.html", encoding="utf-8").read()
if "__DATA__" not in html or "__CSP__" not in html:
    erro("template.html sem os placeholders __DATA__/__CSP__.")
html = html.replace("__DATA__", data_js)

style = re.search(r"<style>(.*?)</style>", html, re.S).group(1)
script = re.search(r"<script>(.*?)</script>", html, re.S).group(1)


def sha(s):
    return base64.b64encode(hashlib.sha256(s.encode("utf-8")).digest()).decode()


csp = (f"default-src 'none'; img-src data:; "
       f"style-src 'sha256-{sha(style)}'; script-src 'sha256-{sha(script)}'; "
       f"base-uri 'none'; form-action 'none'")
html = html.replace("__CSP__", csp)

open("index.html", "w", encoding="utf-8").write(html)
print(f"index.html gerado ({len(html)} bytes). CSP recalculada.")
